import os
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment


new_excel_directory = "C:\\Users\\furkan.cakir\\Desktop\\FurkanPRS\\Kodlar\\Python_Excel_Data_Entry\\Şablonlar_rev01\\yeni_exceller"
old_excel_directory = "C:\\Users\\furkan.cakir\\Desktop\\FurkanPRS\\Kodlar\\Python_Excel_Data_Entry\\Şablonlar_rev01\\old_exceller"
json_file_path = "C:\\Users\\furkan.cakir\\Desktop\\FurkanPRS\\Kodlar\\Python_Excel_Data_Entry\\copy_old_schema_to_new_schema\\excel_names_2.json"

with open(json_file_path, "r", encoding="utf-8") as json_file:
    excel_names_json = json.load(json_file)

new_excel_files = {file for file in os.listdir(
    new_excel_directory) if file.endswith('.xlsx')}
modified_excel_names = ["PROSO KALEM ANA VERİSİ ÖRNEK ŞABLONU_" +
                        name + ".xlsx" for name in excel_names_json]

matching_files = []
found_cells = []

for old_excel_file in os.listdir(old_excel_directory):
    if old_excel_file and not old_excel_file.startswith("~$"):
        old_excel_file_path = os.path.join(old_excel_directory, old_excel_file)

        if os.path.exists(old_excel_file_path):
            processed_sheets = set()

            old_workbook = load_workbook(
                old_excel_file_path, read_only=True, data_only=True)
            variant_index = old_excel_file.lower().find("varyant")
            workbook_prefix = old_excel_file[:variant_index].strip()

            old_sheet_names = {sheet.lower()
                               for sheet in old_workbook.sheetnames}

            for sheet_name in excel_names_json:
                try:
                    sheet_name_lower = sheet_name.lower()
                    if sheet_name_lower in old_sheet_names and (sheet_name not in processed_sheets):
                        sheet_workbook = old_workbook[sheet_name]
                        search_values = ["MODUL", "MODÜL"]
                        for row_index, row in enumerate(sheet_workbook.iter_rows(values_only=True), start=1):
                            kodlar_values = []
                            aciklama_values = []
                            modul_values = []
                            modul_found = False
                            kod_found = False
                            modul_col_index = None

                            for col_index, cell_value in enumerate(row, start=1):
                                if cell_value:
                                    cell_value_lower = str(cell_value).lower()
                                    if "modul" in cell_value_lower or "modül" in cell_value_lower:
                                        modul_found = True
                                        modul_col_index = col_index

                                    elif ("kod" in cell_value_lower or "kodlar" in cell_value_lower) and modul_found:
                                        kod_found = True
                                        break  # No need to check the rest of the row once both conditions are met

                            if modul_found and kod_found:

                                found_cells.append({
                                    "sheet_name": sheet_name,
                                    "cell_value": cell_value,
                                    "row_index": row_index,
                                    "col_index": modul_col_index
                                })

                                for row_index_2 in range(row_index + 1, sheet_workbook.max_row + 1):
                                    kodlar_values.append(sheet_workbook.cell(
                                        row=row_index_2, column=col_index).value)
                                    aciklama_values.append(sheet_workbook.cell(
                                        row=row_index_2, column=col_index + 1).value)
                                    modul_values.append(sheet_workbook.cell(
                                        row=row_index_2, column=modul_col_index).value)

                                # Adjust the order of modul_values to meet your desired format
                                modul_values = modul_values[3:] + \
                                    modul_values[:3]
                                aciklama_values = aciklama_values[3:] + \
                                    aciklama_values[:3]

                                if kodlar_values and aciklama_values and modul_values:
                                    found_cells[-1]["kodlar_values"] = kodlar_values
                                    found_cells[-1]["aciklama_values"] = aciklama_values
                                    found_cells[-1]["modul_values"] = modul_values

                                    # Inside the loop where you iterate over new_excel_files

                                    for new_excel_file in new_excel_files:
                                        split_name = new_excel_file.split(
                                            "_")[1].split(".")[0].lower()
                                        if len(split_name) > 1 and split_name == sheet_name_lower:
                                            new_excel_file_path = os.path.join(
                                                new_excel_directory, new_excel_file)
                                            new_workbook = load_workbook(
                                                new_excel_file_path)
                                            new_sheet = new_workbook.active

                                            # Assuming new_sheet is the worksheet object
                                            columns_to_check = [
                                                1, 2, 4, 5]

                                            # Assuming value, aciklama_values, workbook_prefix, modul_values are defined before this point
                                            for index, value in enumerate(kodlar_values):
                                                # Find the latest used row for each specified column

                                                latest_rows = []

                                                for col in columns_to_check:
                                                    column_values = list(new_sheet.iter_cols(
                                                        min_col=col, max_col=col, values_only=True))
                                                    non_empty_cells = [i for i, cell_value in enumerate(
                                                        column_values[0], start=1) if cell_value is not None]
                                                    latest_row = max(
                                                        non_empty_cells) if non_empty_cells else 0
                                                    latest_rows.append(
                                                        latest_row)

                                                next_row = max(
                                                    latest_rows) + 1

                                                # Now you can use next_row as the row parameter when writing to the sheet
                                                new_sheet.cell(
                                                    row=next_row, column=1, value=value)
                                                new_sheet.cell(
                                                    row=next_row, column=2, value=aciklama_values[index - 3])
                                                new_sheet.cell(
                                                    row=next_row, column=4, value=workbook_prefix).alignment = Alignment(horizontal='center')
                                                new_sheet.cell(
                                                    row=next_row, column=5, value=modul_values[index - 3]).alignment = Alignment(horizontal='center')
                                            # Save the workbook and close it as before
                                            new_workbook.save(
                                                new_excel_file_path)
                                            new_workbook.close()
                                            print(
                                                "yazıldı:", workbook_prefix, "=>", split_name)
                                            processed_sheets.add(
                                                sheet_name)

                                            break  # Break the loop if a match is found
                                        if sheet_name in processed_sheets:
                                            break
                                    if sheet_name in processed_sheets:
                                        break
                                if sheet_name in processed_sheets:
                                    break
                            if sheet_name in processed_sheets:
                                break
                        if sheet_name in processed_sheets:
                            continue

                except Exception as e:
                    print(f"Error opening workbook: {e}")

            # print("found_cells:", found_cells)
