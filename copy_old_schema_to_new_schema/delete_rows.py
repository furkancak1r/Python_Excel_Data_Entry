from openpyxl import load_workbook
import os

excel_directory = "C:\\Users\\furkan.cakir\\Desktop\\FurkanPRS\\Kodlar\\Python_Excel_Data_Entry\\Şablonlar_rev01\\yeni_exceller"

# Get a list of all xlsx files in the excel_directory
excel_files = [file for file in os.listdir(excel_directory) if file.endswith('.xlsx')]

# Iterate over each excel file
for file in excel_files:
    # Load the workbook
    workbook = load_workbook(os.path.join(excel_directory, file))
    
    # Iterate over each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        # Get the sheet
        sheet = workbook[sheet_name]
        
        # Delete rows starting from the third row (index 2)
        sheet.delete_rows(3, sheet.max_row - 2)
        
    # Save the modified workbook
    workbook.save(os.path.join(excel_directory, file))
